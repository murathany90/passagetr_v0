from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sqlite3
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook

SMART_QUOTES_MAP = str.maketrans(
    {
        "\u2018": "'",
        "\u2019": "'",
        "\u201c": '"',
        "\u201d": '"',
        "\u00a0": " ",
        "\ufeff": "",
    }
)

REQUIRED_COLUMNS = ("en_word", "tr_meaning_clean")


@dataclass
class BuildStats:
    rows_read: int = 0
    rows_written: int = 0
    duplicates: int = 0
    invalid_rows: int = 0
    empty_meaning_rows: int = 0


def clean_text(value: str) -> str:
    text = (value or "").translate(SMART_QUOTES_MAP)
    text = text.replace("\r", " ").replace("\n", " ")
    text = re.sub(r"[\u200b\u200c\u200d]", "", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_header(value: str) -> str:
    text = clean_text(value).lower()
    text = text.replace("-", "_").replace(" ", "_")
    text = re.sub(r"_+", "_", text)
    return text.strip("_")


def normalize_word(value: str) -> str:
    text = clean_text(value).lower()
    return re.sub(r"\s+", " ", text).strip()


def normalize_pos(value: str) -> str:
    return normalize_word(value)


def build_search_key(normalized_word: str) -> str:
    text = re.sub(r"[^a-z0-9\s]", " ", normalized_word)
    text = re.sub(r"\s+", " ", text).strip()
    return text or normalized_word


def build_hash(en_word_normalized: str, pos_normalized: str, tr_meaning: str) -> str:
    material = "|".join(
        [
            en_word_normalized.strip(),
            pos_normalized.strip(),
            normalize_word(tr_meaning),
        ]
    )
    return hashlib.sha256(material.encode("utf-8")).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


class DictionaryAssetBuilder:
    def __init__(
        self,
        excel_file: Path,
        output_db: Path,
        sheet_name: Optional[str],
        dataset_version: str,
        batch_size: int,
        report_file: Path,
    ) -> None:
        self.excel_file = excel_file
        self.output_db = output_db
        self.sheet_name = sheet_name
        self.dataset_version = dataset_version
        self.batch_size = batch_size
        self.report_file = report_file

        self.stats = BuildStats()
        self.source_checksum = sha256_file(excel_file)
        self._seen_hashes: set[str] = set()
        self._next_seq_id = 1

    def run(self) -> Dict[str, Any]:
        started = time.perf_counter()

        self.output_db.parent.mkdir(parents=True, exist_ok=True)
        if self.output_db.exists():
            self.output_db.unlink()

        connection = sqlite3.connect(str(self.output_db))
        try:
            self._create_schema(connection)
            rows_buffer: List[tuple] = []

            for row_payload in self._iter_rows():
                rows_buffer.append(row_payload)
                if len(rows_buffer) >= self.batch_size:
                    self._insert_rows(connection, rows_buffer)
                    rows_buffer.clear()

            if rows_buffer:
                self._insert_rows(connection, rows_buffer)

            self._write_bootstrap_meta(connection)
            connection.commit()
            self._optimize(connection)
        finally:
            connection.close()

        elapsed_ms = int((time.perf_counter() - started) * 1000)
        db_size = self.output_db.stat().st_size if self.output_db.exists() else 0

        report = {
            "excel_file": str(self.excel_file),
            "sheet": self.sheet_name,
            "output_db": str(self.output_db),
            "dataset_version": self.dataset_version,
            "source_checksum": self.source_checksum,
            "rows_read": self.stats.rows_read,
            "rows_written": self.stats.rows_written,
            "duplicates": self.stats.duplicates,
            "invalid_rows": self.stats.invalid_rows,
            "empty_meaning_rows": self.stats.empty_meaning_rows,
            "db_size_bytes": db_size,
            "db_size_mb": round(db_size / (1024 * 1024), 2),
            "duration_ms": elapsed_ms,
            "generated_at": datetime.now(timezone.utc).isoformat(),
        }

        self.report_file.parent.mkdir(parents=True, exist_ok=True)
        self.report_file.write_text(
            json.dumps(report, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return report

    def _iter_rows(self) -> Iterable[tuple]:
        workbook = load_workbook(self.excel_file, read_only=True, data_only=True)
        worksheet = self._resolve_worksheet(workbook)
        self.sheet_name = worksheet.title

        try:
            row_iter = worksheet.iter_rows(values_only=True)
            header = next(row_iter, None)
            if not header:
                raise ValueError("Excel dosyasi bos veya header satiri yok.")

            mapping = self._build_column_mapping(header)

            for _, row in enumerate(row_iter, start=2):
                self.stats.rows_read += 1

                en_word = clean_text(self._read_column(row, mapping, "en_word"))
                tr_meaning = clean_text(
                    self._read_column(row, mapping, "tr_meaning_clean")
                )
                pos_clean = clean_text(self._read_column(row, mapping, "pos"))

                if not en_word:
                    self.stats.invalid_rows += 1
                    continue

                if not tr_meaning:
                    self.stats.empty_meaning_rows += 1
                    continue

                normalized = normalize_word(en_word)
                if not normalized:
                    self.stats.invalid_rows += 1
                    continue

                pos_normalized = normalize_pos(pos_clean)
                hash_value = build_hash(normalized, pos_normalized, tr_meaning)

                if hash_value in self._seen_hashes:
                    self.stats.duplicates += 1
                    continue

                self._seen_hashes.add(hash_value)
                search_key = build_search_key(normalized)
                updated_at = int(datetime.now(timezone.utc).timestamp())

                yield (
                    self._next_seq_id,
                    hash_value,
                    en_word,
                    normalized,
                    search_key,
                    pos_clean or None,
                    tr_meaning,
                    "excel_asset",
                    updated_at,
                )
                self._next_seq_id += 1
        finally:
            workbook.close()

    def _create_schema(self, connection: sqlite3.Connection) -> None:
        cursor = connection.cursor()
        cursor.execute("PRAGMA journal_mode = WAL;")
        cursor.execute("PRAGMA synchronous = NORMAL;")
        cursor.execute("PRAGMA temp_store = MEMORY;")

        cursor.executescript(
            """
create table if not exists local_dictionary_entries (
  seq_id integer primary key,
  entry_id text not null,
  en_word text not null,
  en_word_normalized text not null,
  search_key text not null,
  pos text,
  tr_meaning text not null,
  source text not null,
  updated_at integer
);

create unique index if not exists ix_local_dictionary_entry_id
  on local_dictionary_entries (entry_id);

create index if not exists ix_local_dictionary_norm
  on local_dictionary_entries (en_word_normalized);

create index if not exists ix_local_dictionary_search_key
  on local_dictionary_entries (search_key);

create table if not exists local_dictionary_fallback_cache (
  query_normalized text not null,
  query_text text not null,
  source_lang text not null,
  target_lang text not null,
  provider text not null,
  translated_text text not null,
  from_server_cache integer not null default 0,
  hit_count integer not null default 1,
  updated_at integer not null,
  primary key (query_normalized, source_lang, target_lang)
);

create index if not exists ix_local_fallback_updated_at
  on local_dictionary_fallback_cache (updated_at desc);

create table if not exists local_dictionary_bootstrap_meta (
  id integer primary key,
  dataset_version text not null default '',
  batch_id text,
  row_count integer not null default 0,
  downloaded_count integer not null default 0,
  last_seq_id integer not null default 0,
  status text not null default 'idle',
  error_message text,
  updated_at integer not null
);
            """
        )
        cursor.close()

    def _insert_rows(self, connection: sqlite3.Connection, rows: List[tuple]) -> None:
        if not rows:
            return
        cursor = connection.cursor()
        cursor.executemany(
            """
insert or replace into local_dictionary_entries (
  seq_id,
  entry_id,
  en_word,
  en_word_normalized,
  search_key,
  pos,
  tr_meaning,
  source,
  updated_at
)
values (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        self.stats.rows_written += len(rows)
        cursor.close()

    def _write_bootstrap_meta(self, connection: sqlite3.Connection) -> None:
        now_unix = int(datetime.now(timezone.utc).timestamp())
        cursor = connection.cursor()
        cursor.execute("delete from local_dictionary_bootstrap_meta where id = 1")
        cursor.execute(
            """
insert into local_dictionary_bootstrap_meta (
  id,
  dataset_version,
  batch_id,
  row_count,
  downloaded_count,
  last_seq_id,
  status,
  error_message,
  updated_at
)
values (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                1,
                self.dataset_version,
                None,
                self.stats.rows_written,
                self.stats.rows_written,
                max(self.stats.rows_written, 0),
                "ready",
                None,
                now_unix,
            ),
        )
        cursor.close()

    def _optimize(self, connection: sqlite3.Connection) -> None:
        cursor = connection.cursor()
        cursor.execute("ANALYZE;")
        cursor.execute("VACUUM;")
        cursor.close()

    def _resolve_worksheet(self, workbook):
        if not self.sheet_name:
            return workbook.active

        requested = self.sheet_name.strip()
        if requested in workbook.sheetnames:
            return workbook[requested]

        by_casefold = {name.casefold(): name for name in workbook.sheetnames}
        if requested.casefold() in by_casefold:
            return workbook[by_casefold[requested.casefold()]]

        if len(workbook.sheetnames) == 1:
            only_sheet = workbook.sheetnames[0]
            print(
                f"[WARN] Sheet bulunamadi: {requested}. "
                f"Tek sheet bulundu, otomatik secildi: {only_sheet}"
            )
            return workbook[only_sheet]

        available = ", ".join(workbook.sheetnames)
        raise ValueError(
            f"Sheet bulunamadi: {requested}. Mevcut sheetler: {available}"
        )

    def _build_column_mapping(self, header: Iterable[Any]) -> Dict[str, int]:
        mapping: Dict[str, int] = {}
        for idx, raw in enumerate(header):
            key = normalize_header(str(raw or ""))
            if key:
                mapping[key] = idx

        missing = [col for col in REQUIRED_COLUMNS if col not in mapping]
        if missing:
            raise ValueError(
                "Excel header eksik zorunlu kolonlar: " + ", ".join(missing)
            )
        return mapping

    @staticmethod
    def _read_column(row: Iterable[Any], mapping: Dict[str, int], key: str) -> str:
        if key not in mapping:
            return ""
        index = mapping[key]
        row_values = list(row)
        if index >= len(row_values):
            return ""
        value = row_values[index]
        if value is None:
            return ""
        return str(value)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Excel sozluk verisini prebuilt SQLite asset dosyasina donusturur."
    )
    parser.add_argument("--excel-file", required=True, help="Excel dosyasi (.xlsx)")
    parser.add_argument("--sheet", default="", help="Sheet adi (opsiyonel)")
    parser.add_argument(
        "--output-db",
        default="assets/db/dictionary_local.sqlite",
        help="Uretilecek SQLite dosya yolu",
    )
    parser.add_argument(
        "--dataset-version",
        default=datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S"),
        help="Asset dataset versiyonu",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=5000,
        help="SQLite toplu insert batch boyutu",
    )
    parser.add_argument(
        "--report-file",
        default="json_output/dictionary_asset_build_report.json",
        help="JSON rapor dosyasi",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()

    excel_file = Path(args.excel_file).expanduser().resolve()
    if not excel_file.exists():
        print(f"[ERROR] Excel dosyasi bulunamadi: {excel_file}")
        return 1

    if excel_file.suffix.lower() != ".xlsx":
        print("[ERROR] Yalnizca .xlsx dosyalari desteklenir.")
        return 1

    output_db = Path(args.output_db).expanduser().resolve()
    report_file = Path(args.report_file).expanduser().resolve()

    if args.batch_size <= 0:
        print("[ERROR] --batch-size pozitif olmalidir.")
        return 1

    builder = DictionaryAssetBuilder(
        excel_file=excel_file,
        output_db=output_db,
        sheet_name=args.sheet.strip() or None,
        dataset_version=args.dataset_version.strip(),
        batch_size=args.batch_size,
        report_file=report_file,
    )

    try:
        report = builder.run()
    except Exception as err:  # noqa: BLE001
        print(f"[ERROR] Asset build basarisiz: {err}")
        return 1

    print("[OK] Dictionary asset build tamamlandi")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
