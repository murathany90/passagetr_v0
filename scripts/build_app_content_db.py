from __future__ import annotations

import argparse
import hashlib
import json
import sqlite3
import sys
import time
import traceback
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))
LEGACY_SCRIPTS_DIR = ROOT_DIR / "scripts" / "legacy"
if LEGACY_SCRIPTS_DIR.exists() and str(LEGACY_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(LEGACY_SCRIPTS_DIR))

from markdown_to_json_converter import load_grammar_modules_from_markdown
from static_content_common import (
    build_search_key,
    canonicalize_pos,
    clean_text,
    dictionary_entry_id_for_hash,
    normalize_header,
    normalize_word,
    pack_id_for_name,
    passage_id_for_title,
    read_csv_rows,
    read_field,
    sentence_id_for_title_idx,
    sha256_file,
    word_id_for_values,
)


REQUIRED_DICTIONARY_COLUMNS = ("en_word", "tr_meaning_clean")
REQUIRED_WORD_COLUMNS = ("en_word", "tr_meaning", "pos")
REQUIRED_PASSAGE_COLUMNS = ("pack_name", "title", "level", "tags_raw", "category")
REQUIRED_SENTENCE_COLUMNS = ("passage_title", "idx", "sentence_en")


@dataclass
class DictionaryStats:
    rows_read: int = 0
    rows_loaded: int = 0
    duplicate_rows: int = 0
    invalid_rows: int = 0
    empty_meaning_rows: int = 0


@dataclass
class WordsStats:
    rows_read: int = 0
    rows_loaded: int = 0
    duplicate_rows: int = 0
    invalid_rows: int = 0
    empty_meaning_rows: int = 0
    invalid_pos_rows: int = 0


@dataclass
class PassagesStats:
    rows_read: int = 0
    rows_loaded: int = 0
    invalid_rows: int = 0


@dataclass
class SentencesStats:
    rows_read: int = 0
    rows_loaded: int = 0
    invalid_rows: int = 0
    skipped_empty_rows: int = 0
    unmapped_passage_rows: int = 0
    reindexed_rows: int = 0


@dataclass
class GrammarStats:
    module_count: int = 0
    page_count: int = 0
    example_count: int = 0
    test_count: int = 0


class AppContentBuilder:
    def __init__(
        self,
        dictionary_xlsx: Path,
        words_file: Path,
        passages_file: Path,
        sentences_file: Path,
        grammar_dir: Path,
        output_db: Path,
        report_file: Path,
        dataset_version: str,
        skip_grammar: bool,
        word_pack_map_file: Optional[Path],
    ) -> None:
        self.dictionary_xlsx = dictionary_xlsx
        self.words_file = words_file
        self.passages_file = passages_file
        self.sentences_file = sentences_file
        self.grammar_dir = grammar_dir
        self.output_db = output_db
        self.report_file = report_file
        self.dataset_version = dataset_version
        self.skip_grammar = skip_grammar
        self.word_pack_map_file = word_pack_map_file

        self.dictionary_stats = DictionaryStats()
        self.words_stats = WordsStats()
        self.passages_stats = PassagesStats()
        self.sentences_stats = SentencesStats()
        self.grammar_stats = GrammarStats()
        self.grammar_checksum = ""

    def run(self) -> Dict[str, Any]:
        started = time.perf_counter()
        now_unix = int(datetime.now(timezone.utc).timestamp())

        dictionary_rows = self._read_dictionary_rows(now_unix)

        words_mapping, words_rows = read_csv_rows(self.words_file)
        passages_mapping, passages_rows = read_csv_rows(self.passages_file)
        sentences_mapping, sentences_rows = read_csv_rows(self.sentences_file)

        self._ensure_required_headers(
            words_mapping, REQUIRED_WORD_COLUMNS, "YDS_Set_001.csv"
        )
        self._ensure_required_headers(
            passages_mapping, REQUIRED_PASSAGE_COLUMNS, "readings_passages.csv"
        )
        self._ensure_required_headers(
            sentences_mapping, REQUIRED_SENTENCE_COLUMNS, "readings_sentences.csv"
        )

        word_pack_targets = self._load_word_pack_targets()
        pack_ids = self._build_packs(
            passages_rows,
            passages_mapping,
            extra_pack_names=set(word_pack_targets.values()),
        )
        words_payload = self._build_words(
            words_rows,
            words_mapping,
            pack_ids,
            word_pack_targets,
        )
        passages_payload, passage_id_by_title = self._build_passages(
            passages_rows, passages_mapping, pack_ids
        )
        sentences_payload = self._build_sentences(
            sentences_rows, sentences_mapping, passage_id_by_title
        )
        grammar_payload = self._build_grammar_payload() if not self.skip_grammar else {
            "modules": [],
            "pages": [],
            "examples": [],
            "tests": [],
        }

        self.output_db.parent.mkdir(parents=True, exist_ok=True)
        if self.output_db.exists():
            self.output_db.unlink()

        connection = sqlite3.connect(str(self.output_db))
        try:
            self._create_schema(connection)
            self._insert_meta(connection)
            self._insert_packs(connection, pack_ids)
            self._insert_words(connection, words_payload)
            self._insert_passages(connection, passages_payload)
            self._insert_sentences(connection, sentences_payload)
            self._insert_dictionary(connection, dictionary_rows)
            if not self.skip_grammar:
                self._insert_grammar_modules(connection, grammar_payload["modules"])
                self._insert_grammar_pages(connection, grammar_payload["pages"])
                self._insert_grammar_examples(connection, grammar_payload["examples"])
                self._insert_grammar_tests(connection, grammar_payload["tests"])
            self._rebuild_dictionary_fts(connection)
            connection.commit()
            self._optimize(connection)
        finally:
            connection.close()

        elapsed_ms = int((time.perf_counter() - started) * 1000)
        db_size = self.output_db.stat().st_size if self.output_db.exists() else 0
        report = {
            "dataset_version": self.dataset_version,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "duration_ms": elapsed_ms,
            "output_db": str(self.output_db),
            "db_size_bytes": db_size,
            "db_size_mb": round(db_size / (1024 * 1024), 2),
            "checksums": {
                "dictionary_sha256": sha256_file(self.dictionary_xlsx),
                "words_sha256": sha256_file(self.words_file),
                "passages_sha256": sha256_file(self.passages_file),
                "sentences_sha256": sha256_file(self.sentences_file),
                "grammar_sha256": self.grammar_checksum,
            },
            "counts": {
                "packs": len(pack_ids),
                "words": self.words_stats.rows_loaded,
                "reading_passages": self.passages_stats.rows_loaded,
                "reading_sentences": self.sentences_stats.rows_loaded,
                "dictionary_entries": self.dictionary_stats.rows_loaded,
                "dictionary_fts_rows": self.dictionary_stats.rows_loaded,
                "grammar_modules": self.grammar_stats.module_count,
                "grammar_pages": self.grammar_stats.page_count,
                "grammar_examples": self.grammar_stats.example_count,
                "grammar_tests": self.grammar_stats.test_count,
            },
            "stats": {
                "dictionary": self.dictionary_stats.__dict__,
                "words": self.words_stats.__dict__,
                "passages": self.passages_stats.__dict__,
                "sentences": self.sentences_stats.__dict__,
                "grammar": self.grammar_stats.__dict__,
            },
            "word_pack_strategy": {
                "source_id_pack_name": "YDS Set 001",
                "mapping_file": str(self.word_pack_map_file)
                if self.word_pack_map_file is not None
                else None,
                "mapped_word_count": len(word_pack_targets),
                "mapped_pack_names": sorted(set(word_pack_targets.values())),
            },
        }

        meta_sidecar = self.output_db.with_name(f"{self.output_db.stem}.meta.json")
        meta_sidecar.write_text(
            json.dumps(
                {
                    "dataset_version": self.dataset_version,
                    "generated_at": report["generated_at"],
                    "db_file": self.output_db.name,
                    "checksums": report["checksums"],
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        report["meta_sidecar"] = str(meta_sidecar)

        self.report_file.parent.mkdir(parents=True, exist_ok=True)
        self.report_file.write_text(
            json.dumps(report, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return report

    def _read_dictionary_rows(self, now_unix: int) -> List[Tuple[Any, ...]]:
        workbook = load_workbook(self.dictionary_xlsx, read_only=True, data_only=True)
        worksheet = workbook.active
        rows: List[Tuple[Any, ...]] = []
        seen_hashes: set[str] = set()
        next_seq_id = 1

        try:
            row_iter = worksheet.iter_rows(values_only=True)
            header = next(row_iter, None)
            if not header:
                raise ValueError("dictionary.xlsx bos veya header yok.")
            mapping: Dict[str, int] = {}
            for idx, raw in enumerate(header):
                key = normalize_header(str(raw or ""))
                if key:
                    mapping[key] = idx

            missing = [col for col in REQUIRED_DICTIONARY_COLUMNS if col not in mapping]
            if missing:
                raise ValueError(
                    "dictionary.xlsx eksik kolonlar: " + ", ".join(missing)
                )

            for row in row_iter:
                self.dictionary_stats.rows_read += 1
                row_values = [str(cell or "") for cell in row]
                en_word = clean_text(read_field(row_values, mapping, "en_word"))
                tr_meaning = clean_text(read_field(row_values, mapping, "tr_meaning_clean"))
                pos = clean_text(read_field(row_values, mapping, "pos"))

                if not en_word:
                    self.dictionary_stats.invalid_rows += 1
                    continue
                if not tr_meaning:
                    self.dictionary_stats.empty_meaning_rows += 1
                    continue

                normalized = normalize_word(en_word)
                if not normalized:
                    self.dictionary_stats.invalid_rows += 1
                    continue

                material = f"{normalized}|{normalize_word(pos)}|{normalize_word(tr_meaning)}"
                hash_value = hashlib.sha256(material.encode("utf-8")).hexdigest()
                if hash_value in seen_hashes:
                    self.dictionary_stats.duplicate_rows += 1
                    continue
                seen_hashes.add(hash_value)

                entry_id = dictionary_entry_id_for_hash(hash_value)
                search_key = build_search_key(normalized)
                rows.append(
                    (
                        next_seq_id,
                        entry_id,
                        en_word,
                        normalized,
                        search_key,
                        pos or None,
                        tr_meaning,
                        "excel_asset",
                        now_unix,
                    )
                )
                next_seq_id += 1
                self.dictionary_stats.rows_loaded += 1
        finally:
            workbook.close()
        return rows

    def _build_packs(
        self,
        rows: List[List[str]],
        mapping: Dict[str, int],
        extra_pack_names: Optional[set[str]] = None,
    ) -> Dict[str, str]:
        names: set[str] = {"YDS Set 001"}
        for row in rows:
            pack_name = clean_text(read_field(row, mapping, "pack_name"))
            if pack_name:
                names.add(pack_name)
        if extra_pack_names:
            names.update(
                pack_name for pack_name in extra_pack_names if clean_text(pack_name)
            )
        result: Dict[str, str] = {}
        for name in sorted(names):
            result[name] = pack_id_for_name(name)
        return result

    def _load_word_pack_targets(self) -> Dict[str, str]:
        if self.word_pack_map_file is None:
            return {}
        if not self.word_pack_map_file.exists():
            raise RuntimeError(
                f"Kelime paket map dosyasi bulunamadi: {self.word_pack_map_file}"
            )

        raw = json.loads(self.word_pack_map_file.read_text(encoding="utf-8"))
        if isinstance(raw, dict) and isinstance(raw.get("items"), list):
            items = raw.get("items") or []
        elif (
            isinstance(raw, dict)
            and isinstance(raw.get("word_pack_reclassification"), dict)
            and isinstance(raw["word_pack_reclassification"].get("items"), list)
        ):
            items = raw["word_pack_reclassification"].get("items") or []
        else:
            raise RuntimeError(
                "Kelime paket map dosya formati gecersiz. "
                "Beklenen: top-level 'items' veya 'word_pack_reclassification.items'."
            )

        targets: Dict[str, str] = {}
        for item in items:
            if not isinstance(item, dict):
                continue
            word_id = clean_text(str(item.get("word_id") or ""))
            target_pack_name = clean_text(str(item.get("target_pack_name") or ""))
            if word_id and target_pack_name:
                targets[word_id] = target_pack_name
        return targets

    def _build_words(
        self,
        rows: List[List[str]],
        mapping: Dict[str, int],
        pack_ids: Dict[str, str],
        word_pack_targets: Dict[str, str],
    ) -> List[Tuple[Any, ...]]:
        payload: List[Tuple[Any, ...]] = []
        dedupe: set[tuple[str, str]] = set()
        source_pack_name = "YDS Set 001"
        if source_pack_name not in pack_ids:
            raise RuntimeError("YDS Set 001 pack bulunamadi.")

        now_unix = int(datetime.now(timezone.utc).timestamp())
        for row in rows:
            self.words_stats.rows_read += 1
            en_word = clean_text(read_field(row, mapping, "en_word"))
            tr_meaning = clean_text(read_field(row, mapping, "tr_meaning"))
            pos_raw = clean_text(read_field(row, mapping, "pos"))
            example_en = clean_text(read_field(row, mapping, "example_en"))
            example_tr = clean_text(read_field(row, mapping, "example_tr"))
            synonyms_raw = clean_text(read_field(row, mapping, "synonyms_raw"))
            antonyms_raw = clean_text(read_field(row, mapping, "antonyms_raw"))
            level = clean_text(read_field(row, mapping, "level"))
            tags_raw = clean_text(read_field(row, mapping, "tags_raw"))
            notes = clean_text(read_field(row, mapping, "notes"))

            if not en_word:
                self.words_stats.invalid_rows += 1
                continue
            if not tr_meaning:
                self.words_stats.empty_meaning_rows += 1
                continue

            pos_canonical, unknown = canonicalize_pos(pos_raw)
            if not pos_canonical:
                self.words_stats.invalid_rows += 1
                self.words_stats.invalid_pos_rows += 1
                if unknown:
                    print(
                        "[WARN] POS normalize edilemedi:",
                        en_word,
                        "| raw=",
                        pos_raw,
                        "| unknown=",
                        ",".join(unknown),
                    )
                continue

            dedupe_key = (normalize_word(en_word), pos_canonical)
            if dedupe_key in dedupe:
                self.words_stats.duplicate_rows += 1
                continue
            dedupe.add(dedupe_key)

            word_id = word_id_for_values(
                source_pack_name,
                en_word,
                pos_canonical,
            )
            target_pack_name = word_pack_targets.get(word_id, source_pack_name)
            pack_id = pack_ids.get(target_pack_name)
            if not pack_id:
                raise RuntimeError(
                    f"Kelime paketi resolve edilemedi: {target_pack_name}"
                )
            payload.append(
                (
                    word_id,
                    pack_id,
                    en_word,
                    tr_meaning,
                    pos_canonical,
                    pos_raw or None,
                    example_en or en_word,
                    example_tr or None,
                    synonyms_raw or None,
                    antonyms_raw or None,
                    level or None,
                    tags_raw or None,
                    notes or None,
                    normalize_word(en_word),
                    build_search_key(normalize_word(en_word)),
                    now_unix,
                )
            )
            self.words_stats.rows_loaded += 1

        return payload

    def _build_passages(
        self,
        rows: List[List[str]],
        mapping: Dict[str, int],
        pack_ids: Dict[str, str],
    ) -> Tuple[List[Tuple[Any, ...]], Dict[str, str]]:
        payload: List[Tuple[Any, ...]] = []
        by_title: Dict[str, str] = {}
        seen_titles: set[str] = set()
        now_unix = int(datetime.now(timezone.utc).timestamp())

        for row in rows:
            self.passages_stats.rows_read += 1
            pack_name = clean_text(read_field(row, mapping, "pack_name"))
            title = clean_text(read_field(row, mapping, "title"))
            level = clean_text(read_field(row, mapping, "level"))
            tags_raw = clean_text(read_field(row, mapping, "tags_raw"))
            category = clean_text(read_field(row, mapping, "category"))

            if not pack_name or not title:
                self.passages_stats.invalid_rows += 1
                continue

            pack_id = pack_ids.get(pack_name)
            if not pack_id:
                self.passages_stats.invalid_rows += 1
                continue

            title_key = normalize_word(title)
            if title_key in seen_titles:
                self.passages_stats.invalid_rows += 1
                continue
            seen_titles.add(title_key)

            passage_id = passage_id_for_title(title_key)
            by_title[title_key] = passage_id
            payload.append(
                (
                    passage_id,
                    pack_id,
                    pack_name,
                    title,
                    level or None,
                    tags_raw or None,
                    category or None,
                    now_unix,
                )
            )
            self.passages_stats.rows_loaded += 1

        return payload, by_title

    def _build_sentences(
        self,
        rows: List[List[str]],
        mapping: Dict[str, int],
        passage_id_by_title: Dict[str, str],
    ) -> List[Tuple[Any, ...]]:
        grouped: Dict[str, List[Dict[str, Any]]] = {}
        now_unix = int(datetime.now(timezone.utc).timestamp())

        for row in rows:
            self.sentences_stats.rows_read += 1
            passage_title = clean_text(read_field(row, mapping, "passage_title"))
            idx_raw = clean_text(read_field(row, mapping, "idx"))
            sentence_en = clean_text(read_field(row, mapping, "sentence_en"))
            sentence_tr = clean_text(read_field(row, mapping, "sentence_tr"))

            if not passage_title and not idx_raw and not sentence_en and not sentence_tr:
                self.sentences_stats.skipped_empty_rows += 1
                continue

            if not passage_title or not idx_raw or not sentence_en:
                self.sentences_stats.invalid_rows += 1
                continue

            try:
                idx_value = int(idx_raw)
            except ValueError:
                self.sentences_stats.invalid_rows += 1
                continue

            grouped.setdefault(passage_title, []).append(
                {
                    "passage_title": passage_title,
                    "idx": idx_value,
                    "sentence_en": sentence_en,
                    "sentence_tr": sentence_tr or None,
                }
            )

        normalized: List[Dict[str, Any]] = []
        for passage_title, items in grouped.items():
            idx_values = [item["idx"] for item in items]
            has_invalid_idx = any(idx <= 0 for idx in idx_values)
            has_duplicate_idx = len(set(idx_values)) != len(idx_values)
            need_reindex = has_invalid_idx or has_duplicate_idx

            if need_reindex:
                self.sentences_stats.reindexed_rows += len(items)
                for new_idx, item in enumerate(items, start=1):
                    next_item = dict(item)
                    next_item["idx"] = new_idx
                    normalized.append(next_item)
            else:
                normalized.extend(items)

        payload: List[Tuple[Any, ...]] = []
        for item in normalized:
            title_key = normalize_word(item["passage_title"])
            passage_id = passage_id_by_title.get(title_key)
            if not passage_id:
                self.sentences_stats.unmapped_passage_rows += 1
                continue

            sentence_id = sentence_id_for_title_idx(
                item["passage_title"],
                item["idx"],
            )
            payload.append(
                (
                    sentence_id,
                    passage_id,
                    item["passage_title"],
                    item["idx"],
                    item["sentence_en"],
                    item["sentence_tr"],
                    now_unix,
                )
            )
            self.sentences_stats.rows_loaded += 1
        return payload

    def _build_grammar_payload(self) -> Dict[str, List[Tuple[Any, ...]]]:
        if not self.grammar_dir.exists():
            raise FileNotFoundError(f"Grammar klasoru bulunamadi: {self.grammar_dir}")

        self.grammar_checksum = sha256_markdown_dir(self.grammar_dir)
        modules = load_grammar_modules_from_markdown(self.grammar_dir)

        now_unix = int(datetime.now(timezone.utc).timestamp())
        module_rows: List[Tuple[Any, ...]] = []
        page_rows: List[Tuple[Any, ...]] = []
        example_rows: List[Tuple[Any, ...]] = []
        test_rows: List[Tuple[Any, ...]] = []

        for module in modules:
            module_id = int(module.get("sira") or 0)
            if module_id <= 0:
                continue

            module_rows.append(
                (
                    module_id,
                    None,
                    module_id,
                    clean_text(str(module.get("baslik") or "")),
                    clean_text(str(module.get("dosya_adi") or "")),
                    int(module.get("toplam_sayfa") or 0),
                    clean_text(str(module.get("icon") or "📘")),
                    clean_text(str(module.get("renk") or "#4776E6")),
                    now_unix,
                )
            )
            self.grammar_stats.module_count += 1

            pages = module.get("sayfalar") or []
            for page in pages:
                page_no = int(page.get("page_number") or 0)
                if page_no <= 0:
                    continue
                page_id = module_id * 1000 + page_no
                page_rows.append(
                    (
                        page_id,
                        module_id,
                        None,
                        page_no,
                        clean_text(str(page.get("title") or f"Sayfa {page_no}")),
                        str(page.get("content_html") or ""),
                        int(page.get("word_count") or 0),
                    )
                )
                self.grammar_stats.page_count += 1

                for idx, example in enumerate(page.get("examples") or []):
                    example_id = page_id * 1000 + idx + 1
                    example_rows.append(
                        (
                            example_id,
                            page_id,
                            idx,
                            clean_text(str(example.get("en") or "")),
                            clean_text(str(example.get("tr") or "")),
                            clean_text(str(example.get("description") or "")),
                        )
                    )
                    self.grammar_stats.example_count += 1

                tests = page.get("mini_tests") or []
                if not tests and page.get("mini_test"):
                    tests = [page.get("mini_test")]

                for idx, test in enumerate(tests):
                    test_id = page_id * 1000 + idx + 501
                    test_rows.append(
                        (
                            test_id,
                            page_id,
                            idx,
                            clean_text(str(test.get("question") or "")),
                            json.dumps(test.get("options") or {}, ensure_ascii=False),
                            clean_text(str(test.get("correct") or "")),
                            clean_text(str(test.get("explanation") or "")),
                        )
                    )
                    self.grammar_stats.test_count += 1

        return {
            "modules": module_rows,
            "pages": page_rows,
            "examples": example_rows,
            "tests": test_rows,
        }

    def _create_schema(self, connection: sqlite3.Connection) -> None:
        cursor = connection.cursor()
        cursor.execute("PRAGMA journal_mode = WAL;")
        cursor.execute("PRAGMA synchronous = NORMAL;")
        cursor.execute("PRAGMA temp_store = MEMORY;")

        cursor.executescript(
            """
create table if not exists meta (
  key text primary key,
  value text not null
);

create table if not exists packs (
  id text primary key,
  name text not null,
  from_lang text not null,
  to_lang text not null
);

create unique index if not exists ix_packs_name on packs (name);

create table if not exists words (
  id text primary key,
  pack_id text not null,
  en_word text not null,
  tr_meaning text not null,
  pos text not null,
  pos_raw text,
  example_en text not null,
  example_tr text,
  synonyms_raw text,
  antonyms_raw text,
  level text,
  tags_raw text,
  notes text,
  en_word_normalized text not null,
  search_key text not null,
  created_at integer not null
);

create index if not exists ix_words_pack_id on words (pack_id);
create index if not exists ix_words_pack_en on words (pack_id, en_word);
create index if not exists ix_words_pack_pos on words (pack_id, pos);
create index if not exists ix_words_search_key on words (search_key);
create index if not exists ix_words_en_normalized on words (en_word_normalized);

create table if not exists reading_passages (
  id text primary key,
  pack_id text,
  pack_name text,
  title text not null,
  level text,
  tags_raw text,
  category text,
  created_at integer not null
);

create index if not exists ix_reading_passages_pack_id on reading_passages (pack_id);
create index if not exists ix_reading_passages_title on reading_passages (title);
create index if not exists ix_reading_passages_category on reading_passages (category);

create table if not exists reading_sentences (
  id text primary key,
  passage_id text not null,
  passage_title text not null,
  idx integer not null,
  sentence_en text not null,
  sentence_tr text,
  created_at integer not null,
  unique (passage_id, idx)
);

create index if not exists ix_reading_sentences_passage_idx
  on reading_sentences (passage_id, idx);

create table if not exists grammar_modules (
  id integer primary key,
  source_module_id integer,
  sira integer not null,
  baslik text not null,
  dosya_adi text not null,
  toplam_sayfa integer not null default 0,
  icon text not null default '📘',
  renk text not null default '#4776E6',
  updated_at integer not null
);

create unique index if not exists ix_grammar_modules_sira
  on grammar_modules (sira);

create table if not exists grammar_pages (
  id integer primary key,
  module_id integer not null,
  source_page_id integer,
  sayfa_no integer not null,
  baslik text not null,
  icerik_html text not null,
  kelime_sayisi integer not null default 0,
  unique (module_id, sayfa_no)
);

create index if not exists ix_grammar_pages_module_id
  on grammar_pages (module_id);

create table if not exists grammar_examples (
  id integer primary key,
  page_id integer not null,
  sira integer not null default 0,
  ingilizce text not null,
  turkce text not null,
  aciklama text not null default '',
  unique (page_id, sira)
);

create index if not exists ix_grammar_examples_page_id
  on grammar_examples (page_id);

create table if not exists grammar_tests (
  id integer primary key,
  page_id integer not null,
  sira integer not null default 0,
  soru text not null,
  secenekler_json text not null default '{}',
  dogru_cevap text not null default '',
  aciklama text not null default '',
  unique (page_id, sira)
);

create index if not exists ix_grammar_tests_page_id
  on grammar_tests (page_id);

create table if not exists dictionary_entries (
  seq_id integer primary key,
  entry_id text not null,
  en_word text not null,
  en_word_normalized text not null,
  search_key text not null,
  pos text,
  tr_meaning text not null,
  source text not null,
  updated_at integer
);

create unique index if not exists ix_dictionary_entry_id
  on dictionary_entries (entry_id);

create index if not exists ix_dictionary_norm
  on dictionary_entries (en_word_normalized);

create index if not exists ix_dictionary_search_key
  on dictionary_entries (search_key);

create virtual table if not exists dictionary_entries_fts
using fts5(
  en_word,
  tr_meaning,
  pos,
  search_key,
  tokenize='unicode61'
);
            """
        )
        cursor.close()

    def _insert_meta(self, connection: sqlite3.Connection) -> None:
        now_iso = datetime.now(timezone.utc).isoformat()
        rows = [
            ("dataset_version", self.dataset_version),
            ("generated_at", now_iso),
            ("dictionary_sha256", sha256_file(self.dictionary_xlsx)),
            ("words_sha256", sha256_file(self.words_file)),
            ("passages_sha256", sha256_file(self.passages_file)),
            ("sentences_sha256", sha256_file(self.sentences_file)),
            ("grammar_sha256", self.grammar_checksum),
        ]
        cursor = connection.cursor()
        cursor.executemany("insert into meta (key, value) values (?, ?)", rows)
        cursor.close()

    def _insert_packs(self, connection: sqlite3.Connection, pack_ids: Dict[str, str]) -> None:
        cursor = connection.cursor()
        rows = [(pack_id, name, "en", "tr") for name, pack_id in sorted(pack_ids.items())]
        cursor.executemany(
            """
insert into packs (id, name, from_lang, to_lang)
values (?, ?, ?, ?)
            """,
            rows,
        )
        cursor.close()

    def _insert_words(self, connection: sqlite3.Connection, rows: List[Tuple[Any, ...]]) -> None:
        cursor = connection.cursor()
        cursor.executemany(
            """
insert into words (
  id,
  pack_id,
  en_word,
  tr_meaning,
  pos,
  pos_raw,
  example_en,
  example_tr,
  synonyms_raw,
  antonyms_raw,
  level,
  tags_raw,
  notes,
  en_word_normalized,
  search_key,
  created_at
)
values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        cursor.close()

    def _insert_passages(
        self, connection: sqlite3.Connection, rows: List[Tuple[Any, ...]]
    ) -> None:
        cursor = connection.cursor()
        cursor.executemany(
            """
insert into reading_passages (
  id,
  pack_id,
  pack_name,
  title,
  level,
  tags_raw,
  category,
  created_at
)
values (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        cursor.close()

    def _insert_sentences(
        self, connection: sqlite3.Connection, rows: List[Tuple[Any, ...]]
    ) -> None:
        cursor = connection.cursor()
        cursor.executemany(
            """
insert into reading_sentences (
  id,
  passage_id,
  passage_title,
  idx,
  sentence_en,
  sentence_tr,
  created_at
)
values (?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        cursor.close()

    def _insert_grammar_modules(
        self, connection: sqlite3.Connection, rows: List[Tuple[Any, ...]]
    ) -> None:
        if not rows:
            return
        cursor = connection.cursor()
        cursor.executemany(
            """
insert into grammar_modules (
  id,
  source_module_id,
  sira,
  baslik,
  dosya_adi,
  toplam_sayfa,
  icon,
  renk,
  updated_at
)
values (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        cursor.close()

    def _insert_grammar_pages(
        self, connection: sqlite3.Connection, rows: List[Tuple[Any, ...]]
    ) -> None:
        if not rows:
            return
        cursor = connection.cursor()
        cursor.executemany(
            """
insert into grammar_pages (
  id,
  module_id,
  source_page_id,
  sayfa_no,
  baslik,
  icerik_html,
  kelime_sayisi
)
values (?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        cursor.close()

    def _insert_grammar_examples(
        self, connection: sqlite3.Connection, rows: List[Tuple[Any, ...]]
    ) -> None:
        if not rows:
            return
        cursor = connection.cursor()
        cursor.executemany(
            """
insert into grammar_examples (
  id,
  page_id,
  sira,
  ingilizce,
  turkce,
  aciklama
)
values (?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        cursor.close()

    def _insert_grammar_tests(
        self, connection: sqlite3.Connection, rows: List[Tuple[Any, ...]]
    ) -> None:
        if not rows:
            return
        cursor = connection.cursor()
        cursor.executemany(
            """
insert into grammar_tests (
  id,
  page_id,
  sira,
  soru,
  secenekler_json,
  dogru_cevap,
  aciklama
)
values (?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        cursor.close()

    def _insert_dictionary(
        self, connection: sqlite3.Connection, rows: List[Tuple[Any, ...]]
    ) -> None:
        cursor = connection.cursor()
        cursor.executemany(
            """
insert into dictionary_entries (
  seq_id,
  entry_id,
  en_word,
  en_word_normalized,
  search_key,
  pos,
  tr_meaning,
  source,
  updated_at
)
values (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        cursor.close()

    def _rebuild_dictionary_fts(self, connection: sqlite3.Connection) -> None:
        cursor = connection.cursor()
        cursor.execute("delete from dictionary_entries_fts;")
        cursor.execute(
            """
insert into dictionary_entries_fts (rowid, en_word, tr_meaning, pos, search_key)
select seq_id, en_word, tr_meaning, coalesce(pos, ''), search_key
from dictionary_entries
            """
        )
        cursor.close()

    def _optimize(self, connection: sqlite3.Connection) -> None:
        cursor = connection.cursor()
        cursor.execute("ANALYZE;")
        cursor.execute("VACUUM;")
        cursor.close()

    def _ensure_required_headers(
        self,
        mapping: Dict[str, int],
        required_headers: Iterable[str],
        label: str,
    ) -> None:
        missing = [header for header in required_headers if normalize_header(header) not in mapping]
        if missing:
            raise ValueError(f"{label} eksik kolonlar: {', '.join(missing)}")


def sha256_markdown_dir(path: Path) -> str:
    digest = hashlib.sha256()
    files = sorted(path.glob("*.md"), key=lambda p: p.name.lower())
    for file in files:
        digest.update(file.name.encode("utf-8"))
        digest.update(sha256_file(file).encode("utf-8"))
    return digest.hexdigest()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Dictionary + Words + Readings verisini tek app_content.db SQLite assetine donusturur."
    )
    parser.add_argument("--dictionary-xlsx", required=True)
    parser.add_argument("--words-file", required=True)
    parser.add_argument("--passages-file", required=True)
    parser.add_argument("--sentences-file", required=True)
    parser.add_argument("--grammar-dir", default="docs/gramer")
    parser.add_argument("--skip-grammar", action="store_true")
    parser.add_argument("--output-db", default="assets/db/app_content.db")
    parser.add_argument(
        "--report-file", default="json_output/app_content_build_report.json"
    )
    parser.add_argument(
        "--word-pack-map-file",
        default="",
        help="Opsiyonel reclassification raporu. Verilirse kelime pack_id dagitimi bu rapora gore hizalanir.",
    )
    parser.add_argument(
        "--dataset-version",
        default=datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S"),
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    dictionary_xlsx = Path(args.dictionary_xlsx).expanduser().resolve()
    words_file = Path(args.words_file).expanduser().resolve()
    passages_file = Path(args.passages_file).expanduser().resolve()
    sentences_file = Path(args.sentences_file).expanduser().resolve()
    grammar_dir = Path(args.grammar_dir).expanduser().resolve()
    output_db = Path(args.output_db).expanduser().resolve()
    report_file = Path(args.report_file).expanduser().resolve()
    word_pack_map_file = (
        Path(args.word_pack_map_file).expanduser().resolve()
        if str(args.word_pack_map_file or "").strip()
        else None
    )

    for path in (dictionary_xlsx, words_file, passages_file, sentences_file):
        if not path.exists():
            print(f"[ERROR] Dosya bulunamadi: {path}")
            return 1
    if word_pack_map_file is not None and not word_pack_map_file.exists():
        print(f"[ERROR] Kelime paket map dosyasi bulunamadi: {word_pack_map_file}")
        return 1
    if not args.skip_grammar and not grammar_dir.exists():
        print(f"[ERROR] Grammar klasoru bulunamadi: {grammar_dir}")
        return 1

    if dictionary_xlsx.suffix.lower() != ".xlsx":
        print("[ERROR] --dictionary-xlsx yalnizca .xlsx destekler.")
        return 1

    builder = AppContentBuilder(
        dictionary_xlsx=dictionary_xlsx,
        words_file=words_file,
        passages_file=passages_file,
        sentences_file=sentences_file,
        grammar_dir=grammar_dir,
        output_db=output_db,
        report_file=report_file,
        dataset_version=str(args.dataset_version).strip(),
        skip_grammar=bool(args.skip_grammar),
        word_pack_map_file=word_pack_map_file,
    )
    try:
        report = builder.run()
    except Exception as err:  # noqa: BLE001
        print(f"[ERROR] app_content.db build basarisiz: {err}")
        traceback.print_exc()
        return 1

    print("[OK] app_content.db build tamamlandi")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
