from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import Client, create_client


REQUIRED_COLUMNS = ("en_word", "tr_meaning_clean")
OPTIONAL_COLUMNS = ("pos",)
SMART_QUOTES_MAP = str.maketrans(
    {
        "\u2018": "'",
        "\u2019": "'",
        "\u201c": '"',
        "\u201d": '"',
        "\u00a0": " ",
        "\ufeff": "",
    }
)


@dataclass
class ImportStats:
    rows_read: int = 0
    rows_loaded: int = 0
    duplicates: int = 0
    invalid_rows: int = 0
    empty_meaning_rows: int = 0
    inserted_rows: int = 0
    updated_rows: int = 0


class DictionaryImporter:
    _existing_hash_query_chunk_size = 40

    def __init__(
        self,
        client: Client,
        excel_file: Path,
        sheet_name: Optional[str],
        mode: str,
        dataset_version: str,
        batch_size: int,
        report_file: Path,
    ) -> None:
        self.client = client
        self.excel_file = excel_file
        self.sheet_name = sheet_name
        self.mode = mode
        self.dataset_version = dataset_version
        self.batch_size = batch_size
        self.report_file = report_file

        self.stats = ImportStats()
        self.batch_id: Optional[str] = None
        self.source_checksum: str = self._sha256_file(excel_file)
        self._seen_hashes: set[str] = set()

    def run(self) -> Dict[str, Any]:
        start = time.perf_counter()
        payload_rows: List[Dict[str, Any]] = []

        try:
            self.batch_id = self._upsert_import_batch(status="running")

            for payload in self._iter_valid_rows():
                payload_rows.append(payload)
                if len(payload_rows) >= self.batch_size:
                    self._flush_batch(payload_rows)
                    payload_rows.clear()

            if payload_rows:
                self._flush_batch(payload_rows)

            if self.mode == "replace":
                self._deactivate_old_batches()

            self._update_batch_summary(status="completed")
        except Exception as err:  # noqa: BLE001
            self._update_batch_summary(status="failed", error_text=str(err))
            raise

        elapsed_ms = int((time.perf_counter() - start) * 1000)
        report = {
            "excel_file": str(self.excel_file),
            "sheet": self.sheet_name,
            "dataset_version": self.dataset_version,
            "batch_id": self.batch_id,
            "mode": self.mode,
            "source_checksum": self.source_checksum,
            "rows_read": self.stats.rows_read,
            "rows_loaded": self.stats.rows_loaded,
            "duplicates": self.stats.duplicates,
            "invalid_rows": self.stats.invalid_rows,
            "empty_meaning_rows": self.stats.empty_meaning_rows,
            "inserted_rows": self.stats.inserted_rows,
            "updated_rows": self.stats.updated_rows,
            "duration_ms": elapsed_ms,
            "generated_at": datetime.now(timezone.utc).isoformat(),
        }

        self.report_file.parent.mkdir(parents=True, exist_ok=True)
        self.report_file.write_text(
            json.dumps(report, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return report

    def _iter_valid_rows(self) -> Iterable[Dict[str, Any]]:
        workbook = load_workbook(self.excel_file, read_only=True, data_only=True)
        worksheet = self._resolve_worksheet(workbook)
        self.sheet_name = worksheet.title

        try:
            rows = worksheet.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                raise ValueError("Excel dosyasi bos veya header satiri yok.")

            mapping = self._build_column_mapping(header)

            for row_index, row in enumerate(rows, start=2):
                self.stats.rows_read += 1

                raw_en_word = self._read_column(row, mapping, "en_word")
                raw_meaning = self._read_column(row, mapping, "tr_meaning_clean")
                raw_pos = self._read_column(row, mapping, "pos")

                en_word = clean_text(raw_en_word)
                tr_meaning = clean_text(raw_meaning)
                pos_clean = clean_text(raw_pos)

                if not en_word:
                    self.stats.invalid_rows += 1
                    continue

                if not tr_meaning:
                    self.stats.empty_meaning_rows += 1
                    continue

                normalized = normalize_word(en_word)
                if not normalized:
                    self.stats.invalid_rows += 1
                    continue

                pos_normalized = normalize_pos(pos_clean)
                hash_value = build_hash(normalized, pos_normalized, tr_meaning)

                if hash_value in self._seen_hashes:
                    self.stats.duplicates += 1
                    continue
                self._seen_hashes.add(hash_value)

                search_key = build_search_key(normalized)
                meaning_short = build_meaning_short(tr_meaning)

                payload = {
                    "en_word": en_word,
                    "en_word_normalized": normalized,
                    "search_key": search_key,
                    "pos": pos_clean or None,
                    "raw_pos": pos_clean or None,
                    "tr_meaning": tr_meaning,
                    "meaning_short": meaning_short,
                    "source": "excel_import",
                    "is_active": True,
                    "import_batch_id": self.batch_id,
                    "hash": hash_value,
                    "metadata": {
                        "excel_row": row_index,
                        "dataset_version": self.dataset_version,
                    },
                }
                yield payload
        finally:
            workbook.close()

    def _flush_batch(self, rows: List[Dict[str, Any]]) -> None:
        if not rows:
            return

        hashes = [str(item["hash"]) for item in rows]
        existing_hashes: set[str] = set()
        for chunk in self._chunks(
            hashes, chunk_size=self._existing_hash_query_chunk_size
        ):
            existing = self._retry(
                lambda chunk=chunk: self.client.table("dictionary_entries")
                .select("hash")
                .in_("hash", chunk)
                .execute(),
                action="existing hash query",
            )
            for item in (existing.data or []):
                if item and item.get("hash"):
                    existing_hashes.add(str(item["hash"]))

        self.stats.updated_rows += len(existing_hashes)
        self.stats.inserted_rows += len(rows) - len(existing_hashes)

        self._retry(
            lambda: self.client.table("dictionary_entries")
            .upsert(rows, on_conflict="hash")
            .execute(),
            action="dictionary_entries upsert",
        )

        self.stats.rows_loaded += len(rows)

    def _resolve_worksheet(self, workbook):
        if not self.sheet_name:
            return workbook.active

        requested = self.sheet_name.strip()
        if requested in workbook.sheetnames:
            return workbook[requested]

        by_casefold = {name.casefold(): name for name in workbook.sheetnames}
        if requested.casefold() in by_casefold:
            return workbook[by_casefold[requested.casefold()]]

        if len(workbook.sheetnames) == 1:
            only_sheet = workbook.sheetnames[0]
            print(
                f"[WARN] Sheet bulunamadi: {requested}. "
                f"Tek sheet bulundu, otomatik secildi: {only_sheet}"
            )
            return workbook[only_sheet]

        available = ", ".join(workbook.sheetnames)
        raise ValueError(
            f"Sheet bulunamadi: {requested}. Mevcut sheetler: {available}"
        )

    def _deactivate_old_batches(self) -> None:
        if not self.batch_id:
            return

        self._retry(
            lambda: self.client.table("dictionary_entries")
            .update({"is_active": False})
            .neq("import_batch_id", self.batch_id)
            .eq("is_active", True)
            .execute(),
            action="deactivate old dictionary rows",
        )

    def _upsert_import_batch(self, status: str) -> str:
        started_at = datetime.now(timezone.utc).isoformat()
        payload = {
            "dataset_version": self.dataset_version,
            "source_file": str(self.excel_file),
            "source_checksum": self.source_checksum,
            "status": status,
            "started_at": started_at,
            "metadata": {
                "mode": self.mode,
                "sheet": self.sheet_name,
                "batch_size": self.batch_size,
            },
        }

        result = self._retry(
            lambda: self.client.table("dictionary_import_batches")
            .upsert(payload, on_conflict="dataset_version")
            .execute(),
            action="dictionary_import_batches upsert",
        )

        if result.data:
            return str(result.data[0]["id"])

        lookup = self._retry(
            lambda: self.client.table("dictionary_import_batches")
            .select("id")
            .eq("dataset_version", self.dataset_version)
            .limit(1)
            .execute(),
            action="dictionary_import_batches lookup",
        )
        if not lookup.data:
            raise RuntimeError("Import batch ID bulunamadi.")
        return str(lookup.data[0]["id"])

    def _update_batch_summary(self, status: str, error_text: str = "") -> None:
        if not self.batch_id:
            return

        metadata: Dict[str, Any] = {
            "mode": self.mode,
            "sheet": self.sheet_name,
            "batch_size": self.batch_size,
        }
        if error_text.strip():
            metadata["error"] = error_text.strip()[:500]

        payload = {
            "status": status,
            "total_rows_read": self.stats.rows_read,
            "inserted_rows": self.stats.inserted_rows,
            "updated_rows": self.stats.updated_rows,
            "duplicate_rows": self.stats.duplicates,
            "invalid_rows": self.stats.invalid_rows,
            "empty_meaning_rows": self.stats.empty_meaning_rows,
            "completed_at": datetime.now(timezone.utc).isoformat(),
            "metadata": metadata,
        }

        self._retry(
            lambda: self.client.table("dictionary_import_batches")
            .update(payload)
            .eq("id", self.batch_id)
            .execute(),
            action="dictionary_import_batches summary update",
        )

    def _build_column_mapping(self, header: Iterable[Any]) -> Dict[str, int]:
        mapping: Dict[str, int] = {}
        for idx, raw in enumerate(header):
            key = normalize_header(str(raw or ""))
            if key:
                mapping[key] = idx

        missing = [col for col in REQUIRED_COLUMNS if col not in mapping]
        if missing:
            raise ValueError(
                "Excel header eksik zorunlu kolonlar: " + ", ".join(missing)
            )
        return mapping

    @staticmethod
    def _read_column(row: Iterable[Any], mapping: Dict[str, int], key: str) -> str:
        if key not in mapping:
            return ""
        index = mapping[key]
        row_list = list(row)
        if index >= len(row_list):
            return ""
        value = row_list[index]
        if value is None:
            return ""
        return str(value)

    def _retry(self, fn, action: str, retries: int = 5):
        last_error: Optional[Exception] = None
        for attempt in range(1, retries + 1):
            try:
                return fn()
            except Exception as err:  # noqa: BLE001
                last_error = err
                message = str(err).lower()
                retryable = any(
                    token in message
                    for token in (
                        "520",
                        "timeout",
                        "timed out",
                        "connection",
                        "temporarily",
                        "502",
                        "503",
                        "504",
                        "rate limit",
                    )
                )
                if attempt >= retries or not retryable:
                    break
                sleep_seconds = min(2 ** attempt, 20)
                print(
                    f"[WARN] {action} hatasi (deneme {attempt}/{retries}): {err}. "
                    f"{sleep_seconds}s sonra tekrar deneniyor."
                )
                time.sleep(sleep_seconds)
        if last_error is not None:
            raise last_error
        raise RuntimeError(f"Bilinmeyen hata: {action}")

    @staticmethod
    def _sha256_file(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    @staticmethod
    def _chunks(values: List[str], chunk_size: int) -> Iterable[List[str]]:
        if chunk_size <= 0:
            yield values
            return
        for i in range(0, len(values), chunk_size):
            yield values[i : i + chunk_size]


def clean_text(value: str) -> str:
    text = (value or "").translate(SMART_QUOTES_MAP)
    text = text.replace("\r", " ").replace("\n", " ")
    text = re.sub(r"[\u200b\u200c\u200d]", "", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_header(value: str) -> str:
    text = clean_text(value).lower()
    text = text.replace("-", "_").replace(" ", "_")
    text = re.sub(r"_+", "_", text)
    return text.strip("_")


def normalize_word(value: str) -> str:
    text = clean_text(value).lower()
    return re.sub(r"\s+", " ", text).strip()


def normalize_pos(value: str) -> str:
    return normalize_word(value)


def build_search_key(normalized_word: str) -> str:
    text = re.sub(r"[^a-z0-9\s]", " ", normalized_word)
    text = re.sub(r"\s+", " ", text).strip()
    return text or normalized_word


def build_meaning_short(value: str, max_len: int = 120) -> str:
    text = clean_text(value)
    if len(text) <= max_len:
        return text
    return text[: max_len - 3].rstrip() + "..."


def build_hash(en_word_normalized: str, pos_normalized: str, tr_meaning: str) -> str:
    material = "|".join(
        [
            en_word_normalized.strip(),
            pos_normalized.strip(),
            normalize_word(tr_meaning),
        ]
    )
    return hashlib.sha256(material.encode("utf-8")).hexdigest()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Excel sozluk verisini Supabase'e import eder.")
    parser.add_argument("--excel-file", required=True, help="Excel dosyasi (.xlsx)")
    parser.add_argument("--sheet", default="", help="Sheet adi (bos ise aktif sheet)")
    parser.add_argument(
        "--mode",
        choices=("upsert", "replace"),
        default="replace",
        help="replace modunda eski batch satirlari pasiflenir",
    )
    parser.add_argument(
        "--dataset-version",
        default=datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S"),
        help="Import versiyonu (or: 2026-03-03-v1)",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=1000,
        help="Supabase upsert batch boyutu",
    )
    parser.add_argument(
        "--report-file",
        default="json_output/dictionary_import_report.json",
        help="JSON rapor ciktisi",
    )
    parser.add_argument("--supabase-url", default="", help="Opsiyonel SUPABASE_URL override")
    parser.add_argument(
        "--supabase-key",
        default="",
        help="Opsiyonel SUPABASE_SERVICE_ROLE_KEY override",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()

    excel_file = Path(args.excel_file).expanduser().resolve()
    if not excel_file.exists():
        print(f"[ERROR] Excel dosyasi bulunamadi: {excel_file}")
        return 1

    if excel_file.suffix.lower() != ".xlsx":
        print("[ERROR] Yalnizca .xlsx dosyalari desteklenir.")
        return 1

    if args.batch_size <= 0:
        print("[ERROR] --batch-size pozitif olmalidir.")
        return 1

    load_dotenv(dotenv_path=Path(".env"), override=True, encoding="utf-8-sig")

    def _get_env(name: str) -> str:
        return (os.getenv(name, "") or os.getenv(f"\ufeff{name}", "")).strip()

    url = (args.supabase_url or _get_env("SUPABASE_URL")).strip()
    key = (
        args.supabase_key
        or _get_env("SUPABASE_SERVICE_ROLE_KEY")
        or _get_env("SUPABASE_KEY")
    ).strip()

    if not url or not key:
        print("[ERROR] SUPABASE_URL ve SUPABASE_SERVICE_ROLE_KEY zorunludur.")
        return 1

    try:
        client = create_client(url, key)
    except Exception as err:  # noqa: BLE001
        text = str(err)
        if "Invalid API key" in text:
            print(
                "[ERROR] Supabase key formati gecersiz gorunuyor. "
                "Muhtemel neden: supabase-py eski surum."
            )
            print(
                "[HINT] Su komutu calistirin: "
                "python -m pip install --upgrade -r requirements.txt"
            )
            print(
                "[HINT] Ardindan surumu dogrulayin: "
                "python -c \"import supabase; print(supabase.__version__)\""
            )
        else:
            print(f"[ERROR] Supabase istemcisi olusturulamadi: {err}")
        return 1

    importer = DictionaryImporter(
        client=client,
        excel_file=excel_file,
        sheet_name=args.sheet.strip() or None,
        mode=args.mode,
        dataset_version=args.dataset_version.strip(),
        batch_size=args.batch_size,
        report_file=Path(args.report_file),
    )

    try:
        report = importer.run()
    except Exception as err:  # noqa: BLE001
        print(f"[ERROR] Import basarisiz: {err}")
        return 1

    print("[OK] Dictionary import tamamlandi")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
